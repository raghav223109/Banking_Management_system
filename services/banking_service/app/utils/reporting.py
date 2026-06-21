from fpdf import FPDF
import openpyxl
from io import BytesIO
from datetime import datetime

class ReportGenerator:
    @staticmethod
    def generate_statement_pdf(user_name, account_number, transactions):
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(200, 10, txt="Bank Statement", ln=True, align='C')
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt=f"Customer: {user_name}", ln=True)
        pdf.cell(200, 10, txt=f"Account Number: {account_number}", ln=True)
        pdf.cell(200, 10, txt=f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
        pdf.ln(10)

        # Table Header
        pdf.set_font("Arial", 'B', 10)
        pdf.cell(40, 10, "Date", 1)
        pdf.cell(60, 10, "Type", 1)
        pdf.cell(30, 10, "Amount", 1)
        pdf.cell(60, 10, "Description", 1)
        pdf.ln()

        pdf.set_font("Arial", size=10)
        for t in transactions:
            pdf.cell(40, 10, str(t.timestamp.strftime('%Y-%m-%d')), 1)
            pdf.cell(60, 10, str(t.transaction_type.value), 1)
            pdf.cell(30, 10, f"{t.amount:.2f}", 1)
            pdf.cell(60, 10, str(t.description)[:25], 1)
            pdf.ln()

        return pdf.output(dest='S').encode('latin-1')

    @staticmethod
    def generate_statement_excel(user_name, account_number, transactions):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Statement"
        
        headers = ["Date", "Type", "Amount", "Description"]
        ws.append(headers)
        
        for t in transactions:
            ws.append([t.timestamp, t.transaction_type.value, t.amount, t.description])
            
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
